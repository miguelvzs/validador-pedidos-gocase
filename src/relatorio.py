"""Geração dos relatórios Excel formatados na pasta output/.

Formatação deliberadamente simples e profissional: cabeçalho destacado,
cores por prioridade, bordas e larguras ajustadas — sem dashboards elaborados.
"""

import logging
import os
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger("relatorio")

# Quantas vezes tentar mover o arquivo final e o intervalo entre tentativas.
# Dá tempo do gestor fechar a planilha aberta no Excel antes de desistir.
_RETRY_TENTATIVAS = 3
_RETRY_ESPERA_SEG = 4


def _mover_com_retry(origem: Path, destino: Path) -> None:
    """Move o arquivo temporário para o destino final, tolerando lock do Excel.

    Se o destino estiver aberto no Excel (PermissionError), avisa e tenta de
    novo algumas vezes; só então falha com mensagem clara. Escrever num
    temporário primeiro garante que um relatório antigo nunca fica corrompido
    pela metade quando o alvo está travado.
    """
    for tentativa in range(1, _RETRY_TENTATIVAS + 1):
        try:
            os.replace(origem, destino)
            return
        except PermissionError:
            if tentativa < _RETRY_TENTATIVAS:
                logger.warning(
                    "'%s' parece aberto no Excel. Feche-o. Nova tentativa "
                    "%d/%d em %ds...",
                    destino.name, tentativa, _RETRY_TENTATIVAS, _RETRY_ESPERA_SEG,
                )
                time.sleep(_RETRY_ESPERA_SEG)
            else:
                # Remove o temporário para não deixar lixo e falha com instrução.
                Path(origem).unlink(missing_ok=True)
                raise PermissionError(
                    f"Não consegui atualizar '{destino.name}': o arquivo está "
                    "aberto (provavelmente no Excel). Feche-o e rode novamente."
                )


def _caminho_temp(caminho: Path) -> Path:
    """Caminho temporário irmão do destino, para escrita atômica."""
    return caminho.with_name(caminho.stem + ".tmp" + caminho.suffix)

# --- Paleta e estilos compartilhados ---
COR_CABECALHO = "3B5BDB"
CORES_PRIORIDADE = {
    "URGENTE": "FFE0E0",
    "ALTA": "FFF3E0",
    "NORMAL": "E8F5E9",
    "BAIXA": None,  # sem cor de fundo
}
COR_ZEBRA = "F5F5F5"

FONTE_CABECALHO = Font(bold=True, size=11, color="FFFFFF")
FONTE_MOTIVO = Font(bold=True, color="C92A2A")
PREENCHIMENTO_CABECALHO = PatternFill("solid", fgColor=COR_CABECALHO)

_lado = Side(style="thin", color="D0D0D0")
BORDA_FINA = Border(left=_lado, right=_lado, top=_lado, bottom=_lado)

# Colunas cujos valores são monetários — formatadas com 2 casas decimais.
COLUNAS_MONETARIAS = {"valor_unitario", "valor_total"}


def _ajustar_larguras(ws: Worksheet, df: pd.DataFrame) -> None:
    """Ajusta cada coluna ao maior conteúdo (título ou célula) + margem."""
    for idx, coluna in enumerate(df.columns, start=1):
        # Considera o próprio nome da coluna e o maior valor renderizado.
        max_conteudo = max(
            [len(str(coluna))] + [len(str(v)) for v in df[coluna].tolist()]
        )
        ws.column_dimensions[get_column_letter(idx)].width = max_conteudo + 2


def _estilizar_cabecalho(ws: Worksheet, n_colunas: int) -> None:
    """Aplica fundo azul, texto branco e centralização à primeira linha."""
    for col in range(1, n_colunas + 1):
        celula = ws.cell(row=1, column=col)
        celula.fill = PREENCHIMENTO_CABECALHO
        celula.font = FONTE_CABECALHO
        celula.alignment = Alignment(horizontal="center", vertical="center")
        celula.border = BORDA_FINA


def _formatar_monetario(ws: Worksheet, df: pd.DataFrame) -> None:
    """Aplica formato numérico BR (#,##0.00) às colunas monetárias."""
    for idx, coluna in enumerate(df.columns, start=1):
        if coluna in COLUNAS_MONETARIAS:
            letra = get_column_letter(idx)
            for linha in range(2, len(df) + 2):
                ws[f"{letra}{linha}"].number_format = "#,##0.00"


def gerar_relatorio_validados(df: pd.DataFrame, caminho: Path) -> None:
    """Gera Excel com pedidos válidos, formatados por prioridade."""
    os.makedirs(caminho.parent, exist_ok=True)
    temp = _caminho_temp(caminho)

    with pd.ExcelWriter(temp, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Pedidos Válidos")
        ws = writer.sheets["Pedidos Válidos"]

        _estilizar_cabecalho(ws, len(df.columns))

        # Localiza a coluna de prioridade para colorir cada linha conforme faixa.
        col_prioridade = (
            df.columns.get_loc("prioridade") if "prioridade" in df.columns else None
        )

        for linha_excel in range(2, len(df) + 2):
            prioridade = None
            if col_prioridade is not None:
                prioridade = ws.cell(row=linha_excel, column=col_prioridade + 1).value

            cor = CORES_PRIORIDADE.get(str(prioridade))
            for col in range(1, len(df.columns) + 1):
                celula = ws.cell(row=linha_excel, column=col)
                celula.border = BORDA_FINA
                if cor:
                    celula.fill = PatternFill("solid", fgColor=cor)

        _formatar_monetario(ws, df)
        _ajustar_larguras(ws, df)
        ws.freeze_panes = "A2"  # mantém o cabeçalho visível ao rolar

    _mover_com_retry(temp, caminho)
    logger.info("Relatório de válidos gerado: %s (%d linhas)", caminho, len(df))


def gerar_relatorio_rejeitados(df: pd.DataFrame, caminho: Path) -> None:
    """Gera Excel com pedidos rejeitados e motivos de rejeição."""
    os.makedirs(caminho.parent, exist_ok=True)
    temp = _caminho_temp(caminho)

    with pd.ExcelWriter(temp, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Pedidos Rejeitados")
        ws = writer.sheets["Pedidos Rejeitados"]

        _estilizar_cabecalho(ws, len(df.columns))

        col_motivo = (
            df.columns.get_loc("motivo_rejeicao")
            if "motivo_rejeicao" in df.columns
            else None
        )

        for linha_excel in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                celula = ws.cell(row=linha_excel, column=col)
                celula.border = BORDA_FINA
            # Destaca o motivo em vermelho/bold para leitura rápida do problema.
            if col_motivo is not None:
                ws.cell(row=linha_excel, column=col_motivo + 1).font = FONTE_MOTIVO

        _formatar_monetario(ws, df)
        _ajustar_larguras(ws, df)
        ws.freeze_panes = "A2"

    _mover_com_retry(temp, caminho)
    logger.info("Relatório de rejeitados gerado: %s (%d linhas)", caminho, len(df))


def gerar_resumo_execucao(df_validos: pd.DataFrame, df_rejeitados: pd.DataFrame,
                          df_original: pd.DataFrame, tempo_execucao: float,
                          caminho: Path) -> None:
    """Gera Excel com resumo consolidado da execução."""
    os.makedirs(caminho.parent, exist_ok=True)
    temp = _caminho_temp(caminho)

    total = len(df_original)
    n_validos = len(df_validos)
    n_rejeitados = len(df_rejeitados)
    pct_validos = (n_validos / total * 100) if total else 0.0
    pct_rejeitados = (n_rejeitados / total * 100) if total else 0.0

    # Contagem por prioridade (0 quando a faixa não tem pedidos).
    if "prioridade" in df_validos.columns and not df_validos.empty:
        contagem_prio = df_validos["prioridade"].value_counts()
    else:
        contagem_prio = pd.Series(dtype="int64")

    def _prio(faixa: str) -> int:
        return int(contagem_prio.get(faixa, 0))

    valor_validos = float(df_validos["valor_total"].sum()) if not df_validos.empty else 0.0
    valor_rejeitados = (
        float(df_rejeitados["valor_total"].sum(min_count=1) or 0.0)
        if not df_rejeitados.empty else 0.0
    )

    # Marcador de linha separadora: interpretado depois para fazer o merge.
    SEP = "__SEP__"

    linhas: list[tuple[str, object]] = [
        ("Data e hora da execução", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("Total de pedidos processados", total),
        ("Pedidos válidos", f"{n_validos} ({pct_validos:.1f}%)"),
        ("Pedidos rejeitados", f"{n_rejeitados} ({pct_rejeitados:.1f}%)"),
        (SEP, SEP),
        ("Prioridade URGENTE", _prio("URGENTE")),
        ("Prioridade ALTA", _prio("ALTA")),
        ("Prioridade NORMAL", _prio("NORMAL")),
        ("Prioridade BAIXA", _prio("BAIXA")),
        (SEP, SEP),
        ("Valor total dos pedidos válidos (R$)", f"{valor_validos:,.2f}"),
        ("Valor total dos pedidos rejeitados (R$)", f"{valor_rejeitados:,.2f}"),
        (SEP, SEP),
    ]

    # Pedidos por canal, uma linha por canal presente na base original.
    if "canal" in df_original.columns:
        for canal, qtd in df_original["canal"].value_counts().items():
            linhas.append((f"Canal: {canal}", int(qtd)))
        linhas.append((SEP, SEP))

    linhas.append(("Tempo de execução (segundos)", f"{tempo_execucao:.2f}"))

    with pd.ExcelWriter(temp, engine="openpyxl") as writer:
        # Cria a planilha com o cabeçalho; escrevemos as linhas manualmente
        # para poder mesclar as separadoras.
        cabecalho = pd.DataFrame(columns=["Métrica", "Valor"])
        cabecalho.to_excel(writer, index=False, sheet_name="Resumo da Execução")
        ws = writer.sheets["Resumo da Execução"]

        _estilizar_cabecalho(ws, 2)

        fill_sep = PatternFill("solid", fgColor="E9ECEF")
        fill_zebra = PatternFill("solid", fgColor=COR_ZEBRA)

        linha_excel = 2
        for metrica, valor in linhas:
            if metrica == SEP:
                # Separador: mescla as duas colunas numa faixa cinza clara.
                ws.merge_cells(start_row=linha_excel, start_column=1,
                               end_row=linha_excel, end_column=2)
                for col in (1, 2):
                    celula = ws.cell(row=linha_excel, column=col)
                    celula.fill = fill_sep
                    celula.border = BORDA_FINA
            else:
                c_metrica = ws.cell(row=linha_excel, column=1, value=metrica)
                c_valor = ws.cell(row=linha_excel, column=2, value=valor)
                # Zebra: alterna fundo para facilitar a leitura linha a linha.
                if linha_excel % 2 == 0:
                    c_metrica.fill = fill_zebra
                    c_valor.fill = fill_zebra
                c_metrica.font = Font(bold=True)
                for celula in (c_metrica, c_valor):
                    celula.border = BORDA_FINA
            linha_excel += 1

        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 28

    _mover_com_retry(temp, caminho)
    logger.info("Resumo de execução gerado: %s", caminho)
